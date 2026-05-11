"""Supabase-backed database operations.

Drop-in replacement for the old database.py. All functions take an extra
`user_email` argument so each user only sees/modifies their own data.
"""

import os
import io
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
from supabase import create_client, Client

load_dotenv()

# Try to load Streamlit secrets too (so this works in Streamlit Cloud)
try:
    import streamlit as st
    _streamlit_secrets = dict(st.secrets) if hasattr(st, "secrets") else {}
except Exception:
    _streamlit_secrets = {}


def _get_secret(key: str) -> str:
    """Look up a secret in environment first, then Streamlit secrets."""
    return os.getenv(key) or _streamlit_secrets.get(key, "")


# ---------------------------------------------------------------------------
# Connection (cached singleton)
# ---------------------------------------------------------------------------

_client: Client | None = None


def get_client() -> Client:
    """Return a singleton Supabase client."""
    global _client
    if _client is None:
        url = _get_secret("SUPABASE_URL")
        key = _get_secret("SUPABASE_ANON_KEY")
        if not url or not key:
            raise RuntimeError("SUPABASE_URL or SUPABASE_ANON_KEY missing")
        _client = create_client(url, key)
    return _client


# ---------------------------------------------------------------------------
# Auth helpers (Supabase Google OAuth)
# ---------------------------------------------------------------------------

def get_user_from_token(access_token: str) -> dict:
    """Given a Supabase access_token, fetch the user's profile.

    Returns a dict with keys like 'email', 'id', 'user_metadata', etc.
    Raises if the token is invalid.
    """
    if not access_token:
        raise ValueError("No access token provided")

    url = _get_secret("SUPABASE_URL")
    anon_key = _get_secret("SUPABASE_ANON_KEY")
    if not url or not anon_key:
        raise RuntimeError("SUPABASE_URL or SUPABASE_ANON_KEY missing")

    response = requests.get(
        f"{url}/auth/v1/user",
        headers={
            "apikey": anon_key,
            "Authorization": f"Bearer {access_token}",
        },
        timeout=10,
    )
    response.raise_for_status()
    return response.json()


# ---------------------------------------------------------------------------
# Whitelist check
# ---------------------------------------------------------------------------

def is_allowed_user(email: str) -> bool:
    """Check if the given email is in the allowed_users whitelist."""
    if not email:
        return False
    sb = get_client()
    result = sb.table("allowed_users").select("email").eq("email", email).execute()
    return len(result.data) > 0


# ---------------------------------------------------------------------------
# Insert / Update / Delete (always scoped to user_email)
# ---------------------------------------------------------------------------

CONTACT_FIELDS = [
    "full_name", "designation", "company_name", "service_category",
    "email", "phone_primary", "phone_secondary", "whatsapp", "website",
    "address_line", "area", "city", "state", "country", "postal_code",
    "social_linkedin", "notes", "card_image_path"
]


def insert_contact(user_email: str, data: dict) -> int:
    """Insert a contact for the given user, return new id."""
    sb = get_client()
    record = {k: data.get(k) for k in CONTACT_FIELDS}
    record["user_email"] = user_email
    result = sb.table("contacts").insert(record).execute()
    return result.data[0]["id"]


def update_contact(user_email: str, contact_id: int, data: dict):
    """Update a contact (only if it belongs to user_email)."""
    sb = get_client()
    update_fields = [f for f in CONTACT_FIELDS if f != "card_image_path"]
    record = {k: data.get(k) for k in update_fields}
    sb.table("contacts").update(record) \
        .eq("id", contact_id) \
        .eq("user_email", user_email) \
        .execute()


def delete_contact(user_email: str, contact_id: int):
    """Delete a contact (only if it belongs to user_email)."""
    sb = get_client()
    sb.table("contacts").delete() \
        .eq("id", contact_id) \
        .eq("user_email", user_email) \
        .execute()


# ---------------------------------------------------------------------------
# Search / Read
# ---------------------------------------------------------------------------

def search_contacts(user_email: str, category=None, country=None,
                    city=None, area=None, keyword=None) -> list:
    """Search the user's contacts with optional filters."""
    sb = get_client()
    q = sb.table("contacts").select("*").eq("user_email", user_email)

    if category and category != "All":
        q = q.eq("service_category", category)
    if country and country != "All":
        q = q.eq("country", country)
    if city and city.strip():
        q = q.ilike("city", f"%{city.strip()}%")
    if area and area.strip():
        q = q.ilike("area", f"%{area.strip()}%")
    if keyword and keyword.strip():
        kw = keyword.strip()
        result = q.order("created_at", desc=True).execute()
        kw_lower = kw.lower()
        return [
            r for r in result.data
            if any(
                str(r.get(f) or "").lower().find(kw_lower) >= 0
                for f in ["full_name", "company_name", "email",
                          "phone_primary", "notes"]
            )
        ]

    result = q.order("created_at", desc=True).execute()
    return result.data


def get_all_contacts(user_email: str) -> list:
    return search_contacts(user_email)


def get_count(user_email: str) -> int:
    sb = get_client()
    result = sb.table("contacts").select("id", count="exact") \
        .eq("user_email", user_email).execute()
    return result.count or 0


def get_distinct_values(user_email: str, column: str) -> list:
    """Get unique non-null values for a column (for the user's contacts)."""
    allowed = {"service_category", "country", "city", "area"}
    if column not in allowed:
        return []
    sb = get_client()
    result = sb.table("contacts").select(column) \
        .eq("user_email", user_email).execute()
    values = set()
    for row in result.data:
        v = row.get(column)
        if v and str(v).strip():
            values.add(v)
    return sorted(values)


# ---------------------------------------------------------------------------
# Excel export (in-memory bytes - no local files)
# ---------------------------------------------------------------------------

EXPORT_COLUMNS = [
    ("id", "ID"),
    ("full_name", "Full Name"),
    ("designation", "Designation"),
    ("company_name", "Company Name"),
    ("service_category", "Service Category"),
    ("email", "Email"),
    ("phone_primary", "Phone (Primary)"),
    ("phone_secondary", "Phone (Secondary)"),
    ("whatsapp", "WhatsApp"),
    ("website", "Website"),
    ("address_line", "Address Line"),
    ("area", "Area"),
    ("city", "City"),
    ("state", "State"),
    ("country", "Country"),
    ("postal_code", "Postal Code"),
    ("social_linkedin", "LinkedIn"),
    ("notes", "Notes"),
    ("card_image_path", "Card Image Path"),
    ("created_at", "Created At"),
]


def build_excel_bytes(user_email: str) -> bytes:
    """Build an Excel file in memory and return its bytes for download."""
    contacts = get_all_contacts(user_email)

    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"

    headers = [label for _, label in EXPORT_COLUMNS]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="305496")
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row_idx, contact in enumerate(contacts, start=2):
        for col_idx, (key, _) in enumerate(EXPORT_COLUMNS, start=1):
            value = contact.get(key)
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = "" if value is None else str(value)
            cell.number_format = "@"

    width_map = {
        "ID": 6, "Full Name": 22, "Designation": 22, "Company Name": 30,
        "Service Category": 18, "Email": 28, "Phone (Primary)": 20,
        "Phone (Secondary)": 20, "WhatsApp": 18, "Website": 24,
        "Address Line": 35, "Area": 18, "City": 15, "State": 15,
        "Country": 18, "Postal Code": 12, "LinkedIn": 30, "Notes": 30,
        "Card Image Path": 35, "Created At": 20,
    }
    for col_idx, (_, label) in enumerate(EXPORT_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width_map.get(label, 15)

    ws.freeze_panes = "A2"

    last_col_letter = get_column_letter(len(headers))
    last_row = max(2, len(contacts) + 1)
    ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# Storage helpers (card images)
# ---------------------------------------------------------------------------

BUCKET_NAME = "card-images"


def upload_card_image(user_email: str, image_bytes: bytes,
                      filename: str) -> str:
    """Upload an image to storage. Returns the storage path."""
    sb = get_client()
    safe_email = user_email.replace("@", "_at_").replace(".", "_")
    storage_path = f"{safe_email}/{filename}"

    sb.storage.from_(BUCKET_NAME).upload(
        storage_path,
        image_bytes,
        file_options={"content-type": "image/png", "upsert": "true"}
    )
    return storage_path


def get_card_image_url(storage_path: str, expires_in: int = 3600) -> str:
    """Get a signed URL for a card image (valid for `expires_in` seconds)."""
    if not storage_path:
        return ""
    sb = get_client()
    try:
        result = sb.storage.from_(BUCKET_NAME).create_signed_url(
            storage_path, expires_in
        )
        return result.get("signedURL") or result.get("signed_url") or ""
    except Exception:
        return ""


def delete_card_image(storage_path: str):
    """Delete a card image from storage."""
    if not storage_path:
        return
    sb = get_client()
    try:
        sb.storage.from_(BUCKET_NAME).remove([storage_path])
    except Exception:
        pass


if __name__ == "__main__":
    print("cloud_db module loaded.")
    print(f"Bucket: {BUCKET_NAME}")
    sb = get_client()
    print("Supabase client OK")


# ============================================================
# Authentication functions
# ============================================================
import bcrypt


def get_user(username: str):
    """Fetch a user record from Supabase by username. Returns dict or None."""
    try:
        sb = get_client()
        response = sb.table("users").select("*").eq("username", username).limit(1).execute()
        if response.data and len(response.data) > 0:
            return response.data[0]
        return None
    except Exception as e:
        print(f"get_user error: {e}")
        return None


def verify_password(plain: str, hashed: str) -> bool:
    """Verify a plaintext password against a bcrypt hash."""
    try:
        if isinstance(hashed, str):
            hashed = hashed.encode("utf-8")
        return bcrypt.checkpw(plain.encode("utf-8"), hashed)
    except Exception as e:
        print(f"verify_password error: {e}")
        return False


def update_password(username: str, new_password: str) -> bool:
    """Hash a new password and update it in Supabase. Also clears must_change_password."""
    try:
        sb = get_client()
        new_hash = bcrypt.hashpw(new_password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
        response = sb.table("users").update({
            "password_hash": new_hash,
            "must_change_password": False
        }).eq("username", username).execute()
        return bool(response.data)
    except Exception as e:
        print(f"update_password error: {e}")
        return False
